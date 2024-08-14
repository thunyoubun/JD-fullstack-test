from fastapi import FastAPI, File, UploadFile , Response
from fastapi.responses import FileResponse , JSONResponse
from app.utils import read_excel
from pydantic import BaseModel
from typing import List
from starlette.middleware.cors import CORSMiddleware
from openpyxl import Workbook

import shutil
import os

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],

)

class ExcelData(BaseModel):
    columns: List[str]
    # array of array of any type
    rows: List[List]
    

@app.get("/")
async def root():
    return {"msg": "Hello World"}

@app.get("/read/")
async def read_excel_file():
    df = read_excel("part_file.xlsx")
    df = df.fillna('')
    df = df.replace([float('inf'), -float('inf')], 0)  
    columns = list(df.columns)
    rows = df.to_dict(orient="records")
    return {"columns": columns, "rows": rows}

@app.post("/upload/")
async def upload_excel(file: UploadFile = File(...)):
    try:
        with open("part_file.xlsx", "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"message": "File uploaded and overwritten successfully!"}
    except Exception as e:
        return JSONResponse(content={"message": f"Error: {e}"}, status_code=500)
       
@app.get("/export/")
async def export_excel_file():
    file_path = "part_file.xlsx"

    if os.path.exists(file_path):
        return FileResponse(
            path=file_path, 
            filename="downloaded_file.xlsx", 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        return Response(content="File not found", status_code=404)

@app.post("/update/")
async def update_excel_file(data : ExcelData):
 
    # ทำการสร้างไฟล์ excel ใหม่
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # สร้างหัวตาราง
    for idx, column in enumerate(data.columns):
        ws.cell(row=1, column=idx+1, value=column)
    
    # สร้างข้อมูลในตาราง
    for row_idx, row in enumerate(data.rows):
        for col_idx, value in enumerate(row):
            ws.cell(row=row_idx+2, column=col_idx+1, value=value)
    
    # บันทึกไฟล์
    wb.save("part_file.xlsx")

    return {"message": "File updated successfully!"}
